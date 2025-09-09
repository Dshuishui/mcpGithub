#!/usr/bin/env python3
"""
tools/excel_processor.py
Excel文件处理工具模块
"""

import os
import sys

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def register_excel_tools(server):
    """
    注册Excel处理相关的MCP工具到服务器
    
    参数:
    - server: MCP服务器实例
    """
    @server.tool()
    def smart_column_mapping(source_file: str, target_file: str):
        """
        对比两个Excel文件的列结构，为AI映射分析提供清晰的数据展示
        
        参数:
        - source_file: 源文件路径（要复制数据的文件）
        - target_file: 目标文件路径（要接收数据的文件）
        
        AI使用指导:
        请根据列名和数据样本分析列的对应关系，然后输出JSON格式的映射规则：
        格式: {"源列号": "目标列号", "源列号": "目标列号", ...}
        示例: {"1": "3", "2": "1", "3": "2"} 表示源列1映射到目标列3，源列2映射到目标列1，源列3映射到目标列2
        
        常见映射模式:
        - 文件路径相关: path, file, location → File Path, Component Location
        - 组件名称相关: name, component, package → Component Name, Package Name  
        - 状态分类相关: status, category, type → DLT Status, DLT Category
        
        如果某列无法找到合适映射，请在JSON中省略该列
        """
        if not EXCEL_AVAILABLE:
            return "❌ Excel处理功能不可用"
        
        try:
            # 检查文件存在性
            if not os.path.exists(source_file):
                return f"❌ 源文件不存在: {source_file}"
            if not os.path.exists(target_file):
                return f"❌ 目标文件不存在: {target_file}"
            
            # 分析源文件结构
            source_wb = load_workbook(source_file, read_only=True)
            source_sheet = source_wb.active
            source_headers = []
            source_samples = []
            
            for col in range(1, source_sheet.max_column + 1):
                # 获取列名
                header = source_sheet.cell(row=1, column=col).value
                header = str(header) if header is not None else f"Column_{col}"
                source_headers.append(header)
                
                # 获取该列的6个数据样本
                samples = []
                for row in range(2, min(8, source_sheet.max_row + 1)):
                    cell_value = source_sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        samples.append(str(cell_value))
                source_samples.append(samples)
            
            source_wb.close()
            
            # 分析目标文件结构
            target_wb = load_workbook(target_file, read_only=True)
            target_sheet = target_wb.active
            target_headers = []
            target_samples = []
            
            for col in range(1, target_sheet.max_column + 1):
                # 获取列名
                header = target_sheet.cell(row=1, column=col).value
                header = str(header) if header is not None else f"Column_{col}"
                target_headers.append(header)
                
                # 获取该列的6个数据样本
                samples = []
                for row in range(2, min(8, target_sheet.max_row + 1)):
                    cell_value = target_sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        samples.append(str(cell_value))
                target_samples.append(samples)
            
            target_wb.close()
            
            # 构建AI友好的对比结果
            result = f"""📊 文件列结构对比分析

        🗂️ 源文件: {source_file}
        列数: {len(source_headers)}
        ┌─────┬──────────────────┬────────────────────────────────────┐
        │ 列号 │ 列名              │ 数据样本                            │
        ├─────┼──────────────────┼────────────────────────────────────┤"""

            for i, (header, samples) in enumerate(zip(source_headers, source_samples), 1):
                samples_str = " | ".join(samples[:5]) if samples else "无数据"
                result += f"\n│ {i:2d}  │ {header:<16} │ {samples_str:<34} │"

            result += f"""
        └─────┴──────────────────┴────────────────────────────────────┘

        🗂️ 目标文件: {target_file}  
        列数: {len(target_headers)}
        ┌─────┬──────────────────┬────────────────────────────────────┐
        │ 列号 │ 列名              │ 数据样本                            │
        ├─────┼──────────────────┼────────────────────────────────────┤"""

            for i, (header, samples) in enumerate(zip(target_headers, target_samples), 1):
                samples_str = " | ".join(samples[:5]) if samples else "无数据"
                result += f"\n│ {i:2d}  │ {header:<16} │ {samples_str:<34} │"

            result += f"""
        └─────┴──────────────────┴────────────────────────────────────┘

        🤖 AI映射指导:
        请分析上述列结构，输出映射JSON，格式如: {{"1": "3", "2": "1", "3": "2"}}
        说明: 将源文件的列映射到目标文件的对应列"""

            return result
            
        except Exception as e:
            return f"❌ 智能映射分析时出错: {str(e)}"

    @server.tool()
    def copy_data_by_mapping(source_file: str, target_file: str, mapping_rules: str):
        """
        根据映射关系复制数据
        
        参数:
        - source_file: 源文件路径
        - target_file: 目标文件路径  
        - mapping_rules: 映射规则JSON字符串，格式如：
          '{"1": "3", "2": "1", "3": "2"}'  # 源列1→目标列3, 源列2→目标列1, 源列3→目标列2
        """
        if not EXCEL_AVAILABLE:
            return "❌ Excel处理功能不可用"
        
        try:
            # 检查文件存在性
            if not os.path.exists(source_file):
                return f"❌ 源文件不存在: {source_file}"
            if not os.path.exists(target_file):
                return f"❌ 目标文件不存在: {target_file}"
            
            # 解析映射规则
            import json
            try:
                mapping = json.loads(mapping_rules)
            except json.JSONDecodeError:
                return f"❌ 映射规则格式错误，应为JSON格式: {mapping_rules}"
            
            # 读取源文件数据
            source_wb = load_workbook(source_file, read_only=True)
            source_sheet = source_wb.active
            
            # 获取源文件的所有数据（跳过表头）
            source_data = []
            for row in range(2, source_sheet.max_row + 1):
                row_data = []
                for col in range(1, source_sheet.max_column + 1):
                    cell_value = source_sheet.cell(row=row, column=col).value
                    row_data.append(cell_value if cell_value is not None else "")
                source_data.append(row_data)
            
            source_wb.close()
            
            # 打开目标文件进行编辑
            target_wb = load_workbook(target_file)
            target_sheet = target_wb.active
            
            # 保存目标文件的表头
            target_headers = []
            for col in range(1, target_sheet.max_column + 1):
                header = target_sheet.cell(row=1, column=col).value
                target_headers.append(header)
            
            # 清空目标文件的数据行（保留表头）
            if target_sheet.max_row > 1:
                target_sheet.delete_rows(2, target_sheet.max_row - 1)
            
            # 根据映射关系复制数据
            copied_rows = 0
            for src_row_idx, src_row_data in enumerate(source_data):
                target_row_idx = src_row_idx + 2  # 从第2行开始写入（第1行是表头）
                
                # 根据映射规则复制每一列
                for src_col_str, target_col_str in mapping.items():
                    try:
                        src_col = int(src_col_str) - 1  # 转换为0-based索引
                        target_col = int(target_col_str)  # 1-based索引
                        
                        # 检查索引有效性
                        if 0 <= src_col < len(src_row_data):
                            src_value = src_row_data[src_col]
                            target_sheet.cell(row=target_row_idx, column=target_col, value=src_value)
                        
                    except (ValueError, IndexError) as e:
                        continue  # 跳过无效的映射
                
                copied_rows += 1
            
            # 保存目标文件
            target_wb.save(target_file)
            target_wb.close()
            
            # 构建映射描述
            mapping_desc = []
            for src_col, target_col in mapping.items():
                mapping_desc.append(f"源列{src_col}→目标列{target_col}")
            
            result = f"""✅ 数据复制完成
源文件: {source_file} ({len(source_data)}行数据)
目标文件: {target_file}
复制映射: {', '.join(mapping_desc)}
成功复制: {copied_rows}行数据
目标文件表头: {target_headers}"""
            
            return result
            
        except Exception as e:
            return f"❌ 复制数据时出错: {str(e)}"
        
    @server.tool()
    def compare_excel_files(file1: str, file2: str, key_column: str = "1"):
        """
        对比两个Excel文件的差异，用于AI分析
        
        参数:
        - file1: 第一个文件路径（通常是人工维护的清单）
        - file2: 第二个文件路径（通常是系统扫描结果）
        - key_column: 用于匹配的关键列（默认第1列，1-based索引）
        """
        if not EXCEL_AVAILABLE:
            return "❌ Excel处理功能不可用"
        
        try:
            # 检查文件存在性
            if not os.path.exists(file1):
                return f"❌ 文件1不存在: {file1}"
            if not os.path.exists(file2):
                return f"❌ 文件2不存在: {file2}"
            
            # 读取文件1数据
            wb1 = load_workbook(file1, read_only=True)
            sheet1 = wb1.active
            
            headers1 = []
            for col in range(1, sheet1.max_column + 1):
                header = sheet1.cell(row=1, column=col).value
                headers1.append(str(header) if header is not None else f"Column_{col}")
            
            data1 = {}  # 使用字典，key为关键列的值，value为整行数据
            key_col_idx = int(key_column) - 1  # 转换为0-based索引
            
            for row in range(2, sheet1.max_row + 1):
                row_data = []
                for col in range(1, sheet1.max_column + 1):
                    cell_value = sheet1.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                
                # 使用关键列作为key
                if key_col_idx < len(row_data) and row_data[key_col_idx]:
                    key_value = row_data[key_col_idx]
                    data1[key_value] = row_data
            
            wb1.close()
            
            # 读取文件2数据
            wb2 = load_workbook(file2, read_only=True)
            sheet2 = wb2.active
            
            headers2 = []
            for col in range(1, sheet2.max_column + 1):
                header = sheet2.cell(row=1, column=col).value
                headers2.append(str(header) if header is not None else f"Column_{col}")
            
            data2 = {}
            for row in range(2, sheet2.max_row + 1):
                row_data = []
                for col in range(1, sheet2.max_column + 1):
                    cell_value = sheet2.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                
                # 使用关键列作为key
                if key_col_idx < len(row_data) and row_data[key_col_idx]:
                    key_value = row_data[key_col_idx]
                    data2[key_value] = row_data
            
            wb2.close()
            
            # 分析差异
            keys1 = set(data1.keys())
            keys2 = set(data2.keys())
            
            # 只在文件1中存在（已移除的项目）
            only_in_file1 = keys1 - keys2
            # 只在文件2中存在（新增的项目）
            only_in_file2 = keys2 - keys1
            # 两个文件都存在（可能有变更）
            common_keys = keys1 & keys2
            
            # 检查共同项目的数据变更
            modified_items = []
            for key in common_keys:
                if data1[key] != data2[key]:
                    modified_items.append({
                        'key': key,
                        'file1_data': data1[key],
                        'file2_data': data2[key]
                    })
            
            # 构建AI友好的对比结果
            result = f"""📊 Excel文件对比分析

📁 文件1分析: {file1}
表头: {headers1}
数据行数: {len(data1)}

📁 文件2分析: {file2}
表头: {headers2}
数据行数: {len(data2)}

🔍 差异统计:
• 总计文件1项目: {len(data1)}
• 总计文件2项目: {len(data2)}
• 只在文件1中存在: {len(only_in_file1)} (可能已移除)
• 只在文件2中存在: {len(only_in_file2)} (新发现)
• 两文件共有项目: {len(common_keys)}
• 共有项目中有变更: {len(modified_items)}

📝 详细差异:

🚫 只在文件1中存在 (已移除项目):"""
            
            for item in list(only_in_file1)[:5]:  # 最多显示5个
                result += f"\n  - {item}: {data1[item]}"
            if len(only_in_file1) > 5:
                result += f"\n  ... 还有 {len(only_in_file1) - 5} 个项目"
            
            result += f"\n\n🆕 只在文件2中存在 (新发现项目):"
            for item in list(only_in_file2)[:5]:  # 最多显示5个
                result += f"\n  - {item}: {data2[item]}"
            if len(only_in_file2) > 5:
                result += f"\n  ... 还有 {len(only_in_file2) - 5} 个项目"
            
            result += f"\n\n🔄 数据有变更的项目:"
            for item in modified_items[:3]:  # 最多显示3个
                result += f"\n  - {item['key']}:"
                result += f"\n    文件1: {item['file1_data']}"
                result += f"\n    文件2: {item['file2_data']}"
            if len(modified_items) > 3:
                result += f"\n  ... 还有 {len(modified_items) - 3} 个变更项目"
            
            # 添加AI分析用的结构化数据
            result += f"\n\n🤖 AI分析数据:"
            result += f"\n  关键指标: {{"
            result += f"\n    'removed_count': {len(only_in_file1)},"
            result += f"\n    'new_count': {len(only_in_file2)},"
            result += f"\n    'modified_count': {len(modified_items)},"
            result += f"\n    'unchanged_count': {len(common_keys) - len(modified_items)},"
            result += f"\n    'total_file1': {len(data1)},"
            result += f"\n    'total_file2': {len(data2)}"
            result += f"\n  }}"
            
            return result
            
        except Exception as e:
            return f"❌ 对比文件时出错: {str(e)}"

    print("✅ Excel处理工具已注册", file=sys.stderr)