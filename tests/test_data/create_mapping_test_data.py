#!/usr/bin/env python3
"""
create_mapping_test_data.py
生成用于测试智能列映射的Excel文件
"""

import openpyxl
import os

def create_source_file():
    """创建源文件（模拟OneDrive文件，列顺序刻意不同）"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OneDrive Data"
    
    # 源文件的列顺序：组件名称、文件路径、DLT分类 (刻意与目标文件不同)
    headers = ["Package Name", "Component Location", "DLT Category"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 测试数据 (注意列的顺序)
    test_data = [
        ["React Router", "/src/components/router/index.js", "Open Source"],
        ["Lodash Utility", "/src/utils/lodash-helper.py", "Under Review"],  
        ["Bootstrap CSS", "/src/styles/bootstrap.css", "Approved"],
        ["Axios HTTP Client", "/src/api/http-client.json", "Pending Review"],
        ["Moment.js", "/src/utils/date-formatter.js", "Open Source"],
        ["Chart.js", "/src/components/charts/bar-chart.css", "Approved"]
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 确保目录存在
    os.makedirs("tests/test_data", exist_ok=True)
    
    # 保存文件
    file_path = "tests/test_data/source_onedrive.xlsx"
    wb.save(file_path)
    
    print(f"✅ 源文件已创建: {file_path}")
    print(f"   列结构: {headers}")
    return file_path

def create_target_file():
    """创建目标文件（模拟本地文件）"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Local Scan Results"
    
    # 目标文件的列顺序：文件路径、DLT状态、组件名称
    headers = ["File Path", "DLT Status", "Component Name"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 目标文件的现有数据（部分重叠，部分不同）
    test_data = [
        ["/src/components/router/index.js", "Approved", "React Router"],
        ["/src/utils/lodash-helper.py", "Needs Review", "Lodash Utility"],
        ["/src/styles/bootstrap.css", "Approved", "Bootstrap CSS"],
        ["/src/api/new-endpoint.json", "New", "API Endpoint"],  # 新组件
        ["/src/utils/validator.js", "Approved", "Input Validator"],  # 新组件
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    file_path = "tests/test_data/target_local.xlsx"
    wb.save(file_path)
    
    print(f"✅ 目标文件已创建: {file_path}")
    print(f"   列结构: {headers}")
    return file_path

def create_mapping_test_files():
    """创建完整的映射测试文件集"""
    
    print("🎯 生成智能映射测试数据...")
    print("📝 特点：")
    print("   - 源文件和目标文件列顺序不同")
    print("   - 列名略有差异但含义相同") 
    print("   - 数据有重叠但不完全相同")
    print("   - 测试AI的智能映射能力")
    print()
    
    source_file = create_source_file()
    target_file = create_target_file()
    
    print()
    print("🔍 映射挑战:")
    print("   源文件: [Package Name, Component Location, DLT Category]")
    print("   目标文件: [File Path, DLT Status, Component Name]")
    print("   期望映射:")
    print("     Package Name → Component Name")
    print("     Component Location → File Path") 
    print("     DLT Category → DLT Status")
    print()
    print("🧪 现在可以运行测试:")
    print("   python tests/mcp_test_client.py")
    
    return source_file, target_file

if __name__ == "__main__":
    create_mapping_test_files()