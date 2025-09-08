#!/usr/bin/env python3
"""
create_test_excel.py
创建测试用的Excel文件
"""

import openpyxl
import os

def create_test_excel():
    """创建测试用的Excel文件"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Data"
    
    # 添加表头
    headers = ["File Path", "DLT Status", "Component Name"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 添加测试数据
    test_data = [
        ["/path/to/component1.js", "Approved", "React Component"],
        ["/path/to/component2.py", "Under Review", "Python Module"],
        ["/path/to/component3.css", "Approved", "Style Sheet"],
        ["/path/to/component4.json", "Pending", "Config File"],
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 确保目录存在
    os.makedirs("tests/test_data", exist_ok=True)
    
    # 保存文件
    file_path = "tests/test_data/sample.xlsx"
    wb.save(file_path)
    
    print(f"✅ 测试Excel文件已创建: {file_path}")
    return file_path

if __name__ == "__main__":
    create_test_excel()